from bs4 import BeautifulSoup
import requests
import pymysql
import openpyxl
from openpyxl.styles import Font
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
import time


class Stock:
    def __init__(self, *stock_numbers):
        self.stock_numbers = stock_numbers

    def scrape(self):

        result = list()

        for stock_number in self.stock_numbers:

            response = requests.get(
                "https://tw.stock.yahoo.com/q/q?s=" + stock_number)
            soup = BeautifulSoup(response.text.replace("加到投資組合", ""), "lxml")

            stock_date = soup.find(
                "font", {"class": "tt"}).getText().strip()[-9:]  # 資料日期

            tables = soup.find_all("table")[2]  # 取得網頁中第三個表格
            tds = tables.find_all("td")[0:11]  # 取得表格中1到10格

            result.append((stock_date,) +
                          tuple(td.getText().strip() for td in tds))
        return result

    def save(self, stocks):

        db_settings = {
            "host": "127.0.0.1",
            "port": 3306,
            "user": "root",
            "password": "******",
            "db": "stock",
            "charset": "utf8"
        }

        try:
            conn = pymysql.connect(**db_settings)

            with conn.cursor() as cursor:
                sql = """INSERT INTO market(
                                market_date,
                                stock_name,
                                market_time,
                                final_price,
                                buy_price,
                                sell_price,
                                ups_and_downs,
                                lot,
                                yesterday_price,
                                opening_price,
                                highest_price,
                                lowest_price)
                         VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

                for stock in stocks:
                    cursor.execute(sql, stock)
                conn.commit()

        except Exception as ex:
            print("Exception:", ex)

    def export(self, stocks):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("Yahoo股市", 0)

        response = requests.get(
            "https://tw.stock.yahoo.com/q/q?s=2451")
        soup = BeautifulSoup(response.text, "lxml")

        tables = soup.find_all("table")[2]
        ths = tables.find_all("th")[0:11]
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append(titles)

        for index, stock in enumerate(stocks):
            sheet.append(stock)

            if "△" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='FF0000')
            elif "▽" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='00A600')

        wb.save("yahoostock.xlsx")

    def gsheet(self, stocks):
        scopes = ["https://spreadsheets.google.com/feeds"]

        credentials = ServiceAccountCredentials.from_json_keyfile_name(
            "credentials.json", scopes)

        client = gspread.authorize(credentials)

        sheet = client.open_by_key(
            "YOUR GOOGLE SHEET KEY").sheet1

        response = requests.get(
            "https://tw.stock.yahoo.com/q/q?s=2451")
        soup = BeautifulSoup(response.text, "lxml")

        tables = soup.find_all("table")[2]
        ths = tables.find_all("th")[0:11]
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append_row(titles, 1)

        for stock in stocks:
            sheet.append_row(stock)

    def daily(self, year, month):
        browser = webdriver.Chrome(ChromeDriverManager().install())
        browser.get(
            "https://www.twse.com.tw/zh/page/trading/exchange/STOCK_DAY_AVG.html")

        select_year = Select(browser.find_element_by_name("yy"))
        select_year.select_by_value(year)  # 選擇傳入的年份

        select_month = Select(browser.find_element_by_name("mm"))
        select_month.select_by_value(month)  # 選擇傳入的月份

        stockno = browser.find_element_by_name("stockNo")  # 定位股票代碼輸入框

        result = []
        for stock_number in self.stock_numbers:
            stockno.clear()  # 清空股票代碼輸入框
            stockno.send_keys(stock_number)
            stockno.submit()

            time.sleep(2)

            soup = BeautifulSoup(browser.page_source, "lxml")

            table = soup.find("table", {"id": "report-table"})

            elements = table.find_all(
                "td", {"class": "dt-head-center dt-body-center"})

            data = (stock_number,) + tuple(element.getText()
                                           for element in elements)
            result.append(data)

        print(result)


stock = Stock('2451', '2454', '2369')  # 建立Stock物件
stock.daily("2019", "7")  # 動態爬取指定的年月份中，股票代碼的每日收盤價

# stock.gsheet(stock.scrape())  # 將爬取的股票當日行情資料寫入Google Sheet工作表
# stock.export(stock.scrape())  # 將爬取的股票當日行情資料匯出成Excel檔案
# stock.save(stock.scrape())  # 將爬取的股票當日行情資料存入MySQL資料庫中
